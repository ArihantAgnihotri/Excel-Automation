import openpyxl as xl
from openpyxl.styles import PatternFill


def max_name(surnames):
    max = 0
    name = " "
    length = len(surnames)
    for i in range(length):
        if surnames.count(surnames[i]) > max and surnames[i] != ' ':
            max = surnames.count(surnames[i])
            name = surnames[i]

    return name, max


wb = xl.load_workbook("IXB (Updated).xlsx")
sheet = wb["IXB"]
surnames = []

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 3)

    if cell.value is None:
        cell.value = " "

    surnames.append(cell.value)

# We ha a set and ready surname list

surnames.sort()
print(surnames)

maximum = max_name(surnames)
print(f" The surname {maximum[0]} is the most common surname with {maximum[1]} occurences ")
most_used_name = maximum[0]

fill_pat = PatternFill(patternType='solid', fgColor='FF2E2E')

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    if cell.value == most_used_name:
        cell.fill = fill_pat

wb.save("NewSurname.xlsx")
